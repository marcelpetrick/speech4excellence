import openpyxl
import unittest


class ExcelWriter:
    """Handles writing data to an Excel file using openpyxl."""

    def __init__(self, filename):
        """
        Initializes the ExcelWriter with a filename.

        :param filename: str, the filename of the Excel workbook
        """
        self.filename = filename
        try:
            self.workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active

    def insert_pairs(self, pairs):
        """
        Inserts a list of string pairs into the Excel sheet, each on a new row.

        :param pairs: list of tuples, where each tuple contains two strings (artist, title)
        """
        for row, (artist, title) in enumerate(pairs, start=1):
            self.sheet[f'A{row}'] = artist
            self.sheet[f'B{row}'] = title
        self.workbook.save(self.filename)

    def bundle_as_pairs(self, input_list):
        """
        Bundles a list of strings into pairs. Assumes the list length is even.

        :param input_list: list of str, the input list to be bundled into pairs
        :return: list of tuples, bundled input list
        """
        if len(input_list) % 2 != 0:
            raise ValueError("Input list must contain an even number of elements.")
        return [(input_list[i], input_list[i + 1]) for i in range(0, len(input_list), 2)]

    def insert_text(self, text):
        """
        Splits the text by newline and inserts each line into a new row in the Excel sheet.

        :param text: str, the text to be inserted into the Excel sheet
        """
        lines = text.strip().split('\n')
        for row, line in enumerate(lines, start=1):
            self.sheet.cell(row=row, column=1, value=line)
        self.workbook.save(self.filename)


class TestExcelWriter(unittest.TestCase):
    """Unit tests for the ExcelWriter class."""

    def test_bundle_as_pairs_even(self):
        writer = ExcelWriter('test.xlsx')
        input_list = ['Artist1', 'Title1', 'Artist2', 'Title2']
        expected_output = [('Artist1', 'Title1'), ('Artist2', 'Title2')]
        self.assertEqual(writer.bundle_as_pairs(input_list), expected_output)

    def test_bundle_as_pairs_odd(self):
        writer = ExcelWriter('test.xlsx')
        input_list = ['Artist1', 'Title1', 'Artist2']
        with self.assertRaises(ValueError):
            writer.bundle_as_pairs(input_list)


if __name__ == "__main__":
    # Run unit tests when the script is executed directly.
    unittest.main()
